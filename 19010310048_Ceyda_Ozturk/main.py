from openpyxl import load_workbook
from matplotlib import pyplot as plt
import re
import time
from textblob import TextBlob
import pandas as pd
import googletrans
from googletrans import *


wb = load_workbook('C:\\Users\\ceyda\\Desktop\\deneme.xlsx')
sheet = wb.active
pozitif = 0
nötr = 0
negatif = 0
blist = []  # duygular
alist = []  # yorumlar


for cell in sheet['A']:
    alist.append(str(cell.value))

for i in alist:
    print("Yorum Sütunu: "+i)

def graphic():

    dilimler = [pozitif, negatif, nötr]
    degerler = ["Pozitif", "Negatif", "Nötr"]
    color = ["c", "pink", "y"]
    plt.pie(dilimler,
        labels=degerler,
        colors=color,
        autopct='%1.1f%%')

    plt.title = "Film Yorum Analizleri"
    plt.show()

def get_tokenize(text):
    acronym_each_dot = r"(?:[a-zğçşöüı]\.){2,}"
    acronym_end_dot = r"\b[a-zğçşöüı]{2,3}\."
    suffixes = r"[a-zğçşöüı]{3,}' ?[a-zğçşöüı]{0,3}"
    numbers = r"\d+[.,:\d]+"
    any_word = r"[a-zğçşöüı]+"
    punctuations = r"[a-zğçşöüı]*[.,!?;:]"
    word_regex = "|".join([acronym_each_dot,
                           acronym_end_dot,
                           suffixes,
                           numbers,
                           any_word,
                           punctuations])
    return re.compile("%s" % word_regex, re.I).findall(text)


stopwords = frozenset(["acaba", "ama", "ancak", "arada", "bana", "belki", "ben", "benden",
                       "beni", "bence", "benim", "beri", "bile", "bi", "bin", "bir", "biraz", "biri", "birkez",
                       "biz", "bize", "bizden", "bizim", "böyle", "bu", "buna", "bunda", "bundan",
                       "bunlar", "bunu", "bunun", "burada", "çok", "da", "daha", "dahi", "de", "değil", "defa",
                       "diye", "doksan", "edecek", "eden", "ederek", "edilecek", "ediliyor",
                       "ediyor", "elli", "en", "etmesi", "etti", "fakat", "gibi", "halen", "hangi",
                       "hatta", "hem", "hep", "hepsi", "her", "herhangi", "herkesin", "için", "iki",
                       "ile", "ilgili", "ise", "itibaren", "itibariyle", "kadar", "kendi",
                       "kendilerine", "kendini", "kendisi", "kendisine", "kez", "ki", "kim",
                       "kimden", "kime", "kimi", "kimse", "milyar", "milyon", "mu", "ne",
                       "neden", "nedenle", "nerde", "nerede", "nereye", "niye", "olan",
                       "olarak", "oldu", "olduğu", "olduğunu", "olmuş", "olmak", "olmayan", "olmaz", "olsa", "olsun",
                       "olup", "olur", "olursa", "oluyor", "on", "ona", "ondan", "onlar",
                       "onlardan", "onu", "onun", "oysa", "öyle", "pek", "sadece", "sanki", "sekiz",
                       "sen", "senden", "seni", "senin", "siz", "sizden", "sizi", "sizin", "sonra", "şey", "tam", "tek",
                       "tüm",
                       "var", "var.", "ve", "veya", "ya", "yani", "yapacak", "yapmak", "yapılmış", "yok", "yok.",
                       "yine", "yoksa",
                       "zaten", ".", ",", ":", ";", "?", "!", "i", "o"])


def clean_text():
    cleaned_text = []
    for i in alist:
        i = i.lower()
        i = get_tokenize(i)
        i = [y for y in i if not y in stopwords]
        cleaned_text.append(i)
    #for i in cleaned_text:
         #print(i)
    #print(cleaned_text)
    return cleaned_text

cleaned_text = clean_text()


def create_word_freq(corpus, n):
    word_freq = dict()
    for tokens in cleaned_text:
        for token in tokens:
            if token in word_freq.keys():
                word_freq[token] += 1
            else:
                word_freq[token] = 1
    most_freq = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)

    print(dict(most_freq[:n]))
    return dict(most_freq[:n])


most_freq = create_word_freq(cleaned_text, 30)

pozitif_cumleler = []
negatif_cumleler = []
notr_cumleler = []
degistirilmis_cumleler = []
degistirilmis_cumleler2 = []
translator = Translator()
duygu_table = []
for i in alist:
    translate = translator.translate(i, dest='en')
    time.sleep(0.5) #Çeviriye çok istek atınca zaman aşımına uğruyordu geçikme saniyesi koydum
    degistirilmis_cumleler.append(str(translate.text))

for i in degistirilmis_cumleler:
    blob1 = TextBlob(i)
    degistirilmis_cumleler2.append(blob1)
for i in range(100):
    if(degistirilmis_cumleler2[i].polarity > 0):
         duygu_table.append("pozitif")
    elif(degistirilmis_cumleler2[i].polarity < 0):
        duygu_table.append("negatif")
    else:
        duygu_table.append("nötr")

for i in range (2,100):
    sheet.cell(row=i, column=2, value=duygu_table[i])

wb.save(('C:\\Users\\ceyda\\Desktop\\deneme.xlsx'))


for cell in sheet['B']:  # B sütununda ki verileri okuyoruz
    blist.append(str(cell.value))  # Okuduğumuz verileri hlist dizisine atıyoruz
for i in blist:  # B sütununda ki bütün verileri tek tek geziyoruz
    if (i == "pozitif"):
        pozitif = pozitif + 1

    elif (i == "nötr"):
        nötr = nötr + 1

    else:
         negatif = negatif + 1
    pozitif2 = str(pozitif)
    negatif2 = str(negatif)
    nötr2 = str(nötr)
print("Pozitif Duygu Adeti: " +pozitif2)
print("Negatif Duygu Adeti: " +negatif2)
print("Nötr Duygu Adeti: " +nötr2)
graphic()
wb.close()




